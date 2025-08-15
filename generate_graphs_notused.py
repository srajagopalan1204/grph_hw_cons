#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import argparse, json, os, re, sys
from datetime import datetime
from glob import glob
from typing import Dict, List, Optional, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from zoneinfo import ZoneInfo

_LOG_FILE = "graph_generation_log.txt"
_LOG_FH = None

def _open_logfile():
    global _LOG_FH
    try:
        _LOG_FH = open(_LOG_FILE, "a", encoding="utf-8")
    except Exception:
        _LOG_FH = None

def log(msg: str) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    if _LOG_FH:
        try:
            _LOG_FH.write(line + "\n"); _LOG_FH.flush()
        except Exception:
            pass

TIMESTAMP_RE = re.compile(r"_(\d{2})(\d{2})(\d{4})_(\d{2})_(\d{2})")
def extract_timestamp_from_name(name: str) -> Optional[datetime]:
    m = TIMESTAMP_RE.search(name)
    if not m: return None
    mm, dd, yyyy, HH, MM = map(int, m.groups())
    try: return datetime(yyyy, mm, dd, HH, MM)
    except Exception: return None

def _should_ignore(fname: str, terms: List[str]) -> bool:
    low = fname.lower()
    for t in terms + ["_graph", "_graph_", "_grph", " graph", "grph"]:
        if t.lower() in low:
            return True
    return False

def find_latest_in_dir(dir_path: str, terms: List[str]) -> Optional[str]:
    if not os.path.isdir(dir_path): return None
    cand: List[Tuple[datetime,str]] = []
    for fn in os.listdir(dir_path):
        if not fn.lower().endswith(".xlsx"): continue
        if _should_ignore(fn, terms): continue
        full = os.path.join(dir_path, fn)
        ts = extract_timestamp_from_name(fn) or datetime.fromtimestamp(os.path.getmtime(full))
        cand.append((ts, full))
    if not cand: return None
    cand.sort(key=lambda x: x[0])
    return os.path.abspath(cand[-1][1])

def discover_workbooks(paths: List[str], terms: List[str]) -> List[str]:
    out, expanded = [], []
    for p in paths:
        g = glob(p, recursive=True); expanded.extend(g if g else [p])
    seen = set()
    for p in expanded:
        if p in seen: continue
        seen.add(p)
        if os.path.isdir(p):
            latest = find_latest_in_dir(p, terms)
            if latest: out.append(latest)
        elif os.path.isfile(p) and p.lower().endswith(".xlsx") and not _should_ignore(os.path.basename(p), terms):
            out.append(os.path.abspath(p))
    return list(dict.fromkeys(out))

SECTION11_CANDIDATES_FALLBACK = [
    "Section11_Oper_v-name_LoginDt_lstwk","Section11_Oper_v-name_LoginDt_l",
    "Section11_Oper_v_name_LoginDt_l","Section11_Oper_vname_LoginDt_l","Section11_Oper_LoginDt"
]

def _parse_date_prefix_token(p: str):
    m = re.match(r"^(\d{2})(\d{2})(\d{2,4})$", str(p))
    if not m: return p
    mm, dd, yy = m.groups(); yy = ("20"+yy) if len(yy)==2 else yy
    try: return datetime(int(yy), int(mm), int(dd))
    except Exception: return p

def clean_and_highlight_section11(file_path: str, wb, cfg_section11: Dict) -> Optional[str]:
    sheet_candidates = cfg_section11.get("sheet_candidates", SECTION11_CANDIDATES_FALLBACK)
    out_sheet = cfg_section11.get("output_sheet", "Section11_Clean")
    df, found = None, None
    for s in sheet_candidates:
        try:
            df = pd.read_excel(file_path, sheet_name=s); found = s; break
        except Exception: continue
    if df is None or df.empty:
        log("[INFO] Section 11: no candidate sheet found or it's empty — skipping."); return None

    df = df.dropna(how="all")
    if "Oper" in df.columns and "Vname" in df.columns:
        df = df[~(df["Oper"].astype(str).str.strip().eq("") & df["Vname"].astype(str).str.strip().eq(""))]

    login_cols = [c for c in df.columns if isinstance(c, str) and c.endswith("_Login_Date")]
    lastwk_cols = [c for c in df.columns if isinstance(c, str) and c.endswith("_LastWk")]
    def _prefix(col: str, suffix: str): return col[:-len(suffix)] if isinstance(col, str) and col.endswith(suffix) else None
    login_prefixes_sorted = sorted([p for p in [_prefix(c, "_Login_Date") for c in login_cols] if p], key=_parse_date_prefix_token)
    lastwk_prefixes_sorted = sorted([p for p in [_prefix(c, "_LastWk") for c in lastwk_cols] if p], key=_parse_date_prefix_token)
    login_sorted_cols = [f"{p}_Login_Date" for p in login_prefixes_sorted if f"{p}_Login_Date" in df.columns]
    lastwk_sorted_cols = [f"{p}_LastWk" for p in lastwk_prefixes_sorted if f"{p}_LastWk" in df.columns]

    ordered = []
    for base in ["Oper","Vname"]:
        if base in df.columns: ordered.append(base)
    ordered += login_sorted_cols + lastwk_sorted_cols
    for c in df.columns:
        if c not in ordered: ordered.append(c)
    df2 = df[ordered].copy()

    if out_sheet in wb.sheetnames: del wb[out_sheet]
    ws = wb.create_sheet(out_sheet, 0)
    for r in dataframe_to_rows(df2, index=False, header=True): ws.append(r)

    fill = PatternFill(start_color="00FFF59D", end_color="00FFF59D", fill_type="solid")
    name_to_idx = {ws.cell(row=1, column=j).value: j for j in range(1, ws.max_column+1)}

    if len(login_sorted_cols) >= 2:
        rcol, lcol = name_to_idx[login_sorted_cols[-1]], name_to_idx[login_sorted_cols[-2]]
        for rr in range(2, ws.max_row+1):
            vr, vl = ws.cell(row=rr, column=rcol).value, ws.cell(row=rr, column=lcol).value
            if vr is not None and vl is not None and str(vr) == str(vl):
                ws.cell(row=rr, column=rcol).fill = fill

    if len(lastwk_sorted_cols) >= 2:
        rcol, lcol = name_to_idx[lastwk_sorted_cols[-1]], name_to_idx[lastwk_sorted_cols[-2]]
        for rr in range(2, ws.max_row+1):
            vr, vl = ws.cell(row=rr, column=rcol).value, ws.cell(row=rr, column=lcol).value
            if vr is not None and vl is not None and str(vr) == str(vl):
                ws.cell(row=rr, column=rcol).fill = fill

    log(f"[SUCCESS] Section 11 cleaned from '{found}' -> '{out_sheet}' (clustered, corrected highlighting)")
    return out_sheet

from typing import Optional as _Optional
def _parse_mmddyyyy(s: str) -> _Optional[datetime]:
    try: return datetime.strptime(s, "%m%d%Y")
    except Exception: return None

def gather_y_series_by_suffix(df: pd.DataFrame, y_suffix: str, date_prefix_regex: str = r"^(\d{8})_", sort_order: str = "date_asc", coerce_numeric: bool = True):
    pattern = re.compile(date_prefix_regex)
    matches: List[Tuple[datetime, str]] = []
    for c in df.columns:
        if not isinstance(c, str): continue
        if not c.endswith(y_suffix): continue
        m = pattern.match(c)
        if not m: continue
        dt = _parse_mmddyyyy(m.group(1))
        if not dt: continue
        matches.append((dt, c))
    if not matches: return None, [], []
    matches.sort(key=lambda x: x[0], reverse=(sort_order == "date_desc"))
    y_cols_sorted = [col for dt, col in matches]
    y_labels = [dt.strftime("%m/%d/%Y") for dt, col in matches]
    sub = df[y_cols_sorted].copy()
    if coerce_numeric:
        for c in y_cols_sorted: sub[c] = pd.to_numeric(sub[c], errors="coerce")
    return sub, y_cols_sorted, y_labels

def create_chart_on_sheet(wb, df: pd.DataFrame, output_sheet: str, title: str, x_col: str, y_cols: List[str], chart_type: str, label_rotation: int, label_distance: int, y_display_names=None) -> None:
    original_names = getattr(wb, "_original_sheet_names", set())
    if output_sheet in wb.sheetnames:
        if output_sheet in original_names: wb[output_sheet].title = f"{output_sheet} (orig)"
        else: del wb[output_sheet]
    ws = wb.create_sheet(output_sheet)
    headers = [x_col] + (y_display_names if y_display_names else y_cols)
    ws.append(headers)
    for _, row in df.iterrows(): ws.append([row.get(x_col)] + [row.get(c) for c in y_cols])
    t = (chart_type or "line").lower()
    if t in ("column", "stacked_column"):
        chart = BarChart(); chart.type = "col"; chart.grouping = "stacked" if t == "stacked_column" else "clustered"
    elif t == "bar":
        chart = BarChart(); chart.type = "bar"; chart.grouping = "clustered"
    else:
        chart = LineChart()
    chart.title = title; chart.y_axis.title = "Value"; chart.x_axis.title = x_col; chart.x_axis.label_rotation = label_rotation if label_rotation is not None else -45
    data = Reference(ws, min_col=2, min_row=1, max_col=1+len(y_cols), max_row=len(df)+1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(df)+1)
    chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
    ws.add_chart(chart, "E2")

def process_workbook(file_path: str, cfg: Dict, dryrun: bool = False) -> Optional[str]:
    file_path = os.path.abspath(file_path); base_dir = os.path.dirname(file_path); base_name = os.path.basename(file_path)
    if any(x in base_name.lower() for x in ["_graph", "_grph", " graph", "_graph_"]):
        log(f"[INFO] Skipping output-like file '{base_name}'"); return None
    try: wb = load_workbook(file_path)
    except Exception as e: log(f"[ERROR] Failed to open '{file_path}': {e}"); return None
    wb._original_sheet_names = set(wb.sheetnames)

    section11_cfg = cfg.get("section11", {}); modified_sheets: List[str] = []
    if section11_cfg.get("enabled", True):
        try:
            out_sheet = clean_and_highlight_section11(file_path, wb, section11_cfg)
            if out_sheet: modified_sheets.append(out_sheet)
        except Exception as e: log(f"[WARNING] Section 11 processing failed: {e}")

    graph_sheets: List[str] = []; chart_defs: List[Dict] = []
    for wb_cfg in cfg.get("workbooks", []): chart_defs.extend(wb_cfg.get("charts", []))
    cd = cfg.get("charts_defaults", {})
    date_regex_def = cd.get("date_prefix_regex", r"^(\d{8})_")
    sort_def = cd.get("series_sort", "date_asc")
    coerce_def = cd.get("y_coerce_numeric", True)
    series_mode_def = cd.get("y_series_mode", "gather_by_suffix_dateprefix")

    for ch in chart_defs:
        try:
            sheet = ch.get("sheet"); x_col = ch.get("x_col"); output_sheet = ch.get("output_sheet", f"Graph_{sheet}")
            title = ch.get("title", output_sheet); ctype = ch.get("chart_type", cd.get("chart_type", "line"))
            xrot = ch.get("x_label_rotation", cd.get("x_label_rotation", -45)); xdist = ch.get("x_label_distance", cd.get("x_label_distance", 400))
            if not sheet or not x_col: log(f"[WARNING] Chart spec incomplete: {ch} — skipping."); continue
            try: df = pd.read_excel(file_path, sheet_name=sheet)
            except Exception as e: log(f"[WARNING] Unable to load input sheet '{sheet}' in '{base_name}': {e}"); continue
            if df.empty: log(f"[WARNING] Sheet '{sheet}' is empty — skipping chart '{output_sheet}'."); continue

            y_names = None; y_cols = []
            if "y_suffix" in ch or series_mode_def == "gather_by_suffix_dateprefix":
                y_suffix = ch.get("y_suffix")
                if not y_suffix: log(f"[WARNING] y_series_mode requires 'y_suffix' in chart spec: {ch} — skipping."); continue
                dregex = ch.get("date_prefix_regex", date_regex_def); ssort = ch.get("series_sort", sort_def); coer = ch.get("y_coerce_numeric", coerce_def)
                sub, dyn_cols, labels = gather_y_series_by_suffix(df, y_suffix, dregex, ssort, coer)
                if not dyn_cols: log(f"[WARNING] No columns matched y_suffix '{y_suffix}' in sheet '{sheet}' — skipping chart '{output_sheet}'."); continue
                y_cols = dyn_cols; y_names = labels; plot_df = pd.concat([df[[x_col]].copy(), sub], axis=1)
            else:
                y_cols = ch.get("y_cols") or []; 
                if not y_cols: log(f"[WARNING] No y columns specified for chart '{output_sheet}'."); continue
                missing = [c for c in [x_col] + y_cols if c not in df.columns]
                if missing: log(f"[WARNING] Missing columns {missing} in sheet '{sheet}' — skipping chart '{output_sheet}'."); continue
                plot_df = df[[x_col] + y_cols].copy()

            if y_cols:
                mask = plot_df[y_cols].notna().any(axis=1); plot_df = plot_df[mask]
            if len(plot_df) < cfg.get("rules", {}).get("min_rows_per_sheet", 2):
                log(f"[WARNING] Not enough rows for '{output_sheet}' — skipping."); continue

            if not dryrun: create_chart_on_sheet(wb, plot_df, output_sheet, title, x_col, y_cols, ctype, xrot, xdist, y_display_names=y_names)
            graph_sheets.append(output_sheet); log(f"[SUCCESS] Chart created: {output_sheet} in {base_name}")
        except Exception as e:
            log(f"[ERROR] Failed to create chart '{ch}': {e}")

    desired = graph_sheets + modified_sheets + [s for s in wb._original_sheet_names if s not in graph_sheets + modified_sheets]
    try:
        new_order = [wb[name] for name in desired if name in wb.sheetnames]
        for name in wb.sheetnames:
            if name not in desired: new_order.append(wb[name])
        wb._sheets = new_order
    except Exception as e: log(f"[WARNING] Could not fully reorder sheets: {e}")

    tz = ZoneInfo(cfg.get("output", {}).get("timezone", "America/New_York")); now_local = datetime.now(tz)
    cono = (re.search(r"(Cono\d+)", file_path, re.IGNORECASE).group(1) if re.search(r"(Cono\d+)", file_path, re.IGNORECASE) else "ConoX")
    src_ts = extract_timestamp_from_name(base_name) or now_local
    out_name = f"{cono}_Src_{src_ts.strftime('%d%m%y_%H_%M')}__Grph_{now_local.strftime('%d%m%Y%H%M')}.xlsx"
    out_path = os.path.join(base_dir, out_name)
    if dryrun: log(f"[DRYRUN] Would save to: {out_path}")
    else:
        try: wb.save(out_path); log(f"[SUCCESS] Saved: {out_path}")
        except Exception as e: log(f"[ERROR] Failed to save '{out_path}': {e}")

    stamp = now_local.strftime("%m%d%Y_%H%M"); tag = "[DRYRUN]" if dryrun else "[WRITE]"
    log(f"{tag} RUN={stamp} CONO={cono} SOURCE={base_name} OUTPUT={os.path.basename(out_path)}")
    return out_path

def main():
    parser = argparse.ArgumentParser(description="Generate graphs and Section 11 cleanup for Consolidated reports.")
    parser.add_argument("--config", default="config_grph.json", help="Path to config JSON")
    parser.add_argument("--workbooks", nargs="*", default=["./Cono*/Consolidated_reports"], help="Files/dirs/globs to process")
    parser.add_argument("--dryrun", action="store_true", help="Log actions without writing charts or saving")
    args = parser.parse_args()

    _open_logfile()
    tz = ZoneInfo("America/New_York")
    log(datetime.now(tz).strftime("===== RUN %m%d%Y_%H%M (America/New_York) ====="))

    try:
        with open(args.config, "r", encoding="utf-8") as f: cfg = json.load(f)
    except Exception as e:
        log(f"[ERROR] Unable to read config '{args.config}': {e}"); sys.exit(1)

    ignore_terms = [t.lower() for t in cfg.get("discovery", {}).get("ignore_filename_contains", [])]
    targets = []
    for p in args.workbooks:
        if os.path.isdir(p):
            latest = find_latest_in_dir(p, ignore_terms)
            if latest: targets.append(latest)
        else:
            targets.extend(discover_workbooks([p], ignore_terms))

    if not targets:
        log("[WARNING] No workbooks found to process."); sys.exit(0)

    log(f"[INFO] Processing {len(targets)} workbook(s).")
    for fpath in targets:
        log(f"[INFO] -> {fpath}")
        process_workbook(fpath, cfg, dryrun=args.dryrun)

    log("[INFO] Done.")

if __name__ == "__main__":
    main()
