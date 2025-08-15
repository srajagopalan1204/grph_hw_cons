#!/usr/bin/env python3
"""
gen_grph.py (v7.5)
- Charts: JSON-driven only, grph_ prefix.
- Originals copied (values only). Run_Log last.
- Section 11:
  * Clean
  * Reorder -> ALL *_Login_Date (asc), then ALL *_LastWk (asc)
  * Drop rows where all *_Login_Date cells are blank
  * Sort by Oper ascending
  * Conditional formatting:
      - Rightmost *_Login_Date highlighted if equals previous *_Login_Date (per row)
      - Rightmost *_LastWk highlighted if equals previous *_LastWk (per row)
- Uses Series.map (no pandas FutureWarning).
"""
from __future__ import annotations

import os, re, sys, glob, json
import datetime as dt
from typing import Dict, List, Optional, Tuple, Set

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference, Series
from openpyxl.chart.legend import Legend
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

IGNORES = ["_graph", "_graph_", "_Grph"]

def load_config(path: str = "config_grph.json") -> Dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def _is_ignored(filename: str) -> bool:
    fname = os.path.basename(filename)
    if not fname.lower().endswith(".xlsx"):
        return True
    return any(s.lower() in fname.lower() for s in IGNORES)

def _parse_timestamp_from_name(fname: str):
    m = re.search(r'(\d{8})_(\d{2})_(\d{2})', fname)
    if not m:
        return None
    mmddyyyy, hh, mm = m.group(1), m.group(2), m.group(3)
    try:
        return dt.datetime.strptime(mmddyyyy + hh + mm, "%m%d%Y%H%M")
    except Exception:
        return None

def pick_latest_xlsx(path: str) -> Optional[str]:
    files = [p for p in glob.glob(os.path.join(path, "*.xlsx")) if not _is_ignored(p)]
    if not files:
        return None
    scored: List[Tuple[float, str]] = []
    for p in files:
        t = _parse_timestamp_from_name(os.path.basename(p))
        if t:
            scored.append((t.timestamp(), p))
        else:
            scored.append((os.path.getmtime(p), p))
    scored.sort(key=lambda x: x[0], reverse=True)
    return scored[0][1]

def discover_cono_paths() -> List[str]:
    return sorted(glob.glob("./Cono*/Consolidated_reports"))

def normalize_headers(df: pd.DataFrame) -> pd.DataFrame:
    new_cols = []
    for c in df.columns:
        if isinstance(c, str):
            c2 = re.sub(r'\s+', ' ', c.strip())
        else:
            c2 = c
        new_cols.append(c2)
    df = df.copy()
    df.columns = new_cols
    return df

def clean_section11(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in df.columns:
        df[col] = df[col].map(lambda x: x.strip() if isinstance(x, str) else x)
    df.dropna(axis=0, how="all", inplace=True)
    df.dropna(axis=1, how="all", inplace=True)
    if df.empty:
        return df
    header = list(df.columns)
    mask = ~(df.apply(lambda row: list(row.values) == header, axis=1))
    df = df.loc[mask]
    df.reset_index(drop=True, inplace=True)
    return df

def reorder_section11_columns_keep_all(df: pd.DataFrame, run_log: List[str]) -> Tuple[pd.DataFrame, List[str], List[str]]:
    """Reorder to: Oper, Vname, ALL *_Login_Date ASC, then ALL *_LastWk ASC.
       Returns (df_reordered, login_cols_ordered, lastwk_cols_ordered)."""
    df = df.copy()
    cols = list(df.columns)

    def find_col(name_targets):
        for c in cols:
            s = str(c).strip().lower().replace(" ", "").replace("_", "")
            for t in name_targets:
                if s == t:
                    return c
        return None

    oper_col = find_col(["oper"])
    vname_col = find_col(["vname","vendorname","name"])

    date_map = {}
    pat = re.compile(r'^\s*(\d{8})_(Login_Date|LastWk)\s*$', re.IGNORECASE)
    for c in cols:
        m = pat.match(str(c))
        if not m:
            continue
        d8 = m.group(1)
        kind = m.group(2).lower()
        if d8 not in date_map:
            date_map[d8] = {}
        if kind.startswith("login"):
            date_map[d8]["login"] = c
        else:
            date_map[d8]["lastwk"] = c

    if not date_map:
        run_log.append("[S11] No date-tagged columns matched 'MMDDYYYY_(Login_Date|LastWk)'. Leaving as-is.")
        return df, [], []

    dates_sorted = sorted(date_map.keys(), key=lambda s: dt.datetime.strptime(s, "%m%d%Y"))
    login_cols = [date_map[d].get("login") for d in dates_sorted if date_map[d].get("login")]
    lastwk_cols = [date_map[d].get("lastwk") for d in dates_sorted if date_map[d].get("lastwk")]

    ordered = []
    if oper_col: ordered.append(oper_col)
    if vname_col: ordered.append(vname_col)
    ordered += login_cols + lastwk_cols

    # Append non-date leftovers
    placed = set(ordered)
    for c in cols:
        if c not in placed and not pat.match(str(c)):
            ordered.append(c)

    ordered = [c for c in ordered if c in df.columns]
    run_log.append(f"[S11] Column order -> {ordered[:8]}{'...' if len(ordered)>8 else ''}")
    return df[ordered], login_cols, lastwk_cols

def drop_rows_all_login_blank(df: pd.DataFrame, login_cols: List[str], run_log: List[str]) -> pd.DataFrame:
    if not login_cols:
        return df
    # consider blank if NaN or empty string or only whitespace
    def row_blank_login(r):
        for c in login_cols:
            val = r.get(c, None)
            if pd.notna(val) and str(val).strip() != "":
                return False
        return True
    before = len(df)
    mask = df.apply(row_blank_login, axis=1)
    df2 = df.loc[~mask].copy()
    after = len(df2)
    run_log.append(f"[S11] Dropped {before - after} rows with all Login_Date blank.")
    return df2

def sort_by_oper(df: pd.DataFrame, run_log: List[str]) -> pd.DataFrame:
    if "Oper" in df.columns:
        df2 = df.sort_values(by="Oper", kind="stable").reset_index(drop=True)
        run_log.append("[S11] Sorted by Oper ascending.")
        return df2
    # try case-insensitive match
    for c in df.columns:
        if str(c).strip().lower() == "oper":
            df2 = df.sort_values(by=c, kind="stable").reset_index(drop=True)
            run_log.append(f"[S11] Sorted by '{c}' ascending.")
            return df2
    run_log.append("[S11] Oper column not found for sorting.")
    return df

def add_group_highlight(ws, group_cols: List[str]):
    if len(group_cols) < 2:
        return
    # build header -> column index map (row 1)
    header_to_idx = {}
    for col_idx in range(1, ws.max_column + 1):
        header_to_idx[ws.cell(row=1, column=col_idx).value] = col_idx
    right = header_to_idx.get(group_cols[-1])
    left = header_to_idx.get(group_cols[-2])
    if not right or not left:
        return
    last_letter = get_column_letter(right)
    prev_letter = get_column_letter(left)
    cell_range = f"{last_letter}2:{last_letter}{ws.max_row}"
    formula = f"{last_letter}2={prev_letter}2"
    fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    rule = FormulaRule(formula=[formula], fill=fill, stopIfTrue=False)
    ws.conditional_formatting.add(cell_range, rule)

def apply_two_group_highlights(ws, login_cols: List[str], lastwk_cols: List[str]):
    add_group_highlight(ws, login_cols)
    add_group_highlight(ws, lastwk_cols)

def chart_from_type(chart_type: str):
    ctype = (chart_type or "column").lower()
    if ctype in ("column","bar","stacked_column"):
        c = BarChart()
        if ctype == "bar":
            c.type = "bar"
        if ctype == "stacked_column":
            c.type = "col"; c.grouping = "stacked"
        c.legend = Legend(); return c
    if ctype == "line":
        c = LineChart(); c.legend = Legend(); return c
    if ctype == "scatter":
        c = ScatterChart(); c.style = 13; c.legend = Legend(); return c
    c = BarChart(); c.legend = Legend(); return c

def write_df(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

def safe_sheet_name(wb, base: str) -> str:
    base = re.sub(r'[:\\/?*\[\]]', '_', base)
    if len(base) <= 31 and base not in wb.sheetnames: return base
    stem = base[:28]; name = stem; i = 1
    while name in wb.sheetnames or len(name) > 31:
        suffix = f"_{i}"; name = (stem[:31 - len(suffix)]) + suffix; i += 1
    return name

def gather_y_suffixes(df: pd.DataFrame) -> List[str]:
    suf: Set[str] = set()
    for c in df.columns:
        s = c if isinstance(c, str) else str(c)
        m = re.match(r'^\s*(\d{8})_(\w+?)(?:_[xy])?\s*$', s, re.IGNORECASE)
        if m:
            suf.add(m.group(2))
    return sorted(suf)

def gather_y_cols_for_suffix(df: pd.DataFrame, y_suffix: str) -> List[str]:
    pat = re.compile(r'^\s*\d{8}_' + re.escape(y_suffix) + r'(?:_[xy])?\s*$', re.IGNORECASE)
    return [c for c in df.columns if pat.match(c if isinstance(c, str) else str(c))]

def build_and_insert_chart(wb, df: pd.DataFrame, x_col: str, y_suffix: str, chart_type: str, title: str, sort_by_x: Optional[str], log: List[str]):
    df = normalize_headers(df)
    if str(x_col).lower() == "first_column":
        real_x = df.columns[0]
    else:
        target = re.sub(r'\s+', ' ', (x_col or "").strip()).lower()
        real_x = None
        for c in df.columns:
            if isinstance(c, str) and re.sub(r'\s+', ' ', c.strip()).lower() == target:
                real_x = c; break
        if real_x is None:
            log.append(f"[SKIP] x_col '{x_col}' not found. Columns: {list(df.columns)[:12]}")
            return None, 0
    y_cols = gather_y_cols_for_suffix(df, y_suffix)
    if not y_cols:
        log.append(f"[SKIP] No Y columns for suffix '{y_suffix}'. (Expect headers like MMDDYYYY_{y_suffix})")
        return None, 0
    def date_key(c):
        s = c if isinstance(c, str) else str(c)
        m = re.match(r'^\s*(\d{8})_', s)
        if m:
            try:
                return dt.datetime.strptime(m.group(1), "%m%d%Y")
            except Exception:
                return dt.datetime(1900,1,1)
        return dt.datetime(1900,1,1)
    y_cols.sort(key=date_key)
    df2 = df.copy()
    for c in y_cols: df2[c] = pd.to_numeric(df2[c], errors="coerce")
    keep_cols = [real_x] + y_cols; df2 = df2[keep_cols]
    if sort_by_x in ("asc","desc"): df2 = df2.sort_values(by=real_x, ascending=(sort_by_x=="asc"))
    mask_keep = ~df2[y_cols].isna().all(axis=1); df2 = df2.loc[mask_keep]
    if df2.empty: log.append(f"[SKIP] No usable rows for chart '{title}'."); return None, 0
    def pretty_label(colname: str) -> str:
        s = str(colname).strip()
        m = re.match(r'^\s*(\d{2})(\d{2})(\d{4})_', s)
        return f"{m.group(1)}/{m.group(2)}/{m.group(3)}" if m else s
    display_cols = [real_x] + [pretty_label(c) for c in y_cols]
    df_display = df2.copy(); df_display.columns = display_cols
    gname_raw = f"grph_{re.sub(r'[^A-Za-z0-9_]+','', title)}"; gname = safe_sheet_name(wb, gname_raw)
    ws = wb.create_sheet(gname); write_df(ws, df_display)
    ctype = (chart_type or "column").lower(); cobj = chart_from_type(chart_type); cobj.title = title
    nrows = len(df_display) + 1; ncols = len(display_cols)
    if ctype in ("line","column","bar","stacked_column"):
        data_ref = Reference(ws, min_col=2, min_row=1, max_col=ncols, max_row=nrows)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=nrows)
        cobj.add_data(data_ref, titles_from_data=True); cobj.set_categories(cats_ref)
    else:
        xref = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=nrows)
        for idx in range(2, ncols+1):
            yref = Reference(ws, min_col=idx, min_row=1, max_col=idx, max_row=nrows)
            series = Series(yref, xref, title_from_data=True); cobj.series.append(series)
    ws.add_chart(cobj, "H2"); log.append(f"[OK] grph: {gname} ({ncols-1} series, {nrows-1} rows) X='{real_x}'")
    return gname, ncols-1

def find_sheet_name(sheet_names: List[str], sheet: Optional[str], sheet_prefix: Optional[str]) -> Optional[str]:
    if sheet and sheet in sheet_names: return sheet
    if sheet_prefix:
        for s in sheet_names:
            if s.startswith(sheet_prefix): return s
    return None

def norm11(name: str) -> str:
    return re.sub(r'[^a-z0-9]', '', name.lower())

def is_section11(name: str) -> bool:
    n = norm11(name)
    return (n == "section11") or (n == "sec11") or ("section11" in n) or ("sec11" in n)

def apply_two_group_highlights(ws, login_cols: List[str], lastwk_cols: List[str]):
    add_group_highlight(ws, login_cols)
    add_group_highlight(ws, lastwk_cols)

def main():
    cfg = load_config()
    cono_paths = discover_cono_paths()
    if not cono_paths:
        print("[INFO] No Cono paths found under ./Cono*/Consolidated_reports")
        return

    overall_logs = []

    for path in cono_paths:
        cono = re.search(r'Cono(\d+)', path)
        cono_name = f"Cono{cono.group(1)}" if cono else os.path.basename(os.path.dirname(path))

        src = pick_latest_xlsx(path)
        if not src:
            overall_logs.append(f"[{cono_name}] No source workbook found in {path}.")
            continue

        src_dt = _parse_timestamp_from_name(os.path.basename(src)) or dt.datetime.fromtimestamp(os.path.getmtime(src))
        src_ts_str = src_dt.strftime("%d%m%y_%H_%M")

        print(f"[INFO] Processing workbook for {cono_name}: {src}")

        xls = pd.ExcelFile(src, engine="openpyxl")
        sheet_names = xls.sheet_names

        wb = Workbook()
        wb.remove(wb.active)

        run_log: List[str] = []
        created_graphs = 0

        # Charts (JSON-driven only)
        for ch in cfg.get("charts", []):
            target_sheet = find_sheet_name(sheet_names, ch.get("sheet"), ch.get("sheet_prefix"))
            if not target_sheet:
                run_log.append(f"[SKIP] Sheet not found (sheet={ch.get('sheet')}, sheet_prefix={ch.get('sheet_prefix')}).")
                continue

            df = xls.parse(sheet_name=target_sheet, dtype=str)

            x_col = ch.get("x_col", "first_column")
            y_suffix = ch.get("y_suffix", "*")
            chart_type = ch.get("chart_type", "column")
            sort_by_x = ch.get("sort_by_x")

            if y_suffix == "*":
                suffixes = gather_y_suffixes(df)
                if not suffixes:
                    run_log.append(f"[SKIP] No suffixes found in sheet '{target_sheet}'.")
                    continue
                for suf in suffixes:
                    title = ch.get("title", f"{target_sheet}") + f"_{suf}"
                    gname, nseries = build_and_insert_chart(wb, df, x_col, suf, chart_type, title, sort_by_x, run_log)
                    if gname:
                        created_graphs += 1
            else:
                title = ch.get("title", f"{target_sheet}_{y_suffix}")
                gname, nseries = build_and_insert_chart(wb, df, x_col, y_suffix, chart_type, title, sort_by_x, run_log)
                if gname:
                    created_graphs += 1

        # Section 11: clean + reorder (ALL groups) + drop rows + sort by Oper + dual highlights -> Section11_clean
        s11_sheet = next((s for s in sheet_names if is_section11(s)), None)
        if s11_sheet:
            df11 = xls.parse(sheet_name=s11_sheet)
            df11c = clean_section11(df11)
            df11c, login_cols, lastwk_cols = reorder_section11_columns_keep_all(df11c, run_log)
            df11c = drop_rows_all_login_blank(df11c, login_cols, run_log)
            df11c = sort_by_oper(df11c, run_log)

            ws11 = wb.create_sheet(safe_sheet_name(wb, "Section11_clean"))
            write_df(ws11, df11c)
            apply_two_group_highlights(ws11, login_cols, lastwk_cols)
            run_log.append(f"[OK] Section 11 cleaned, reordered, filtered, sorted, highlighted -> 'Section11_clean' (rows={len(df11c)}).")
        else:
            run_log.append("[INFO] Section 11 sheet not found; skipping Section11_clean.")

        # Originals (values-only)
        for sname in sheet_names:
            dfo = xls.parse(sheet_name=sname)
            wso = wb.create_sheet(safe_sheet_name(wb, f"Original_{re.sub(r'[^A-Za-z0-9_]+','', sname)}"))
            write_df(wso, dfo)

        # Run_Log last
        wslog = wb.create_sheet(safe_sheet_name(wb, "Run_Log"))
        wslog.cell(row=1, column=1, value="Event")
        for i, line in enumerate(run_log, start=2):
            wslog.cell(row=i, column=1, value=line)

        # Save near source
        cfg_dir = (cfg.get("output", {}) or {}).get("directory", "same_as_source")
        out_dir = os.path.dirname(src) if (cfg_dir in ("same_as_source", "", None)) else cfg_dir
        os.makedirs(out_dir, exist_ok=True)
        out_pat = cfg.get("output", {}).get("filename_pattern", "{cono}_Src_{src_ts}__Grph_{run_ts_EST}")
        outname = out_pat.format(cono=cono_name, src_ts=src_ts_str, run_ts_EST=dt.datetime.now().strftime("%m%d%Y_%H%M"))
        out_path = os.path.join(out_dir, f"{outname}.xlsx")

        from openpyxl.writer.excel import save_workbook
        save_workbook(wb, out_path)
        overall_logs.append(f"[{cono_name}] Wrote {out_path} (graphs={created_graphs}).")

    print("\n".join(overall_logs))

if __name__ == "__main__":
    main()
