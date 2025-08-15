import argparse
import json
import re
import sys
from pathlib import Path
from datetime import datetime
from zoneinfo import ZoneInfo
from dateutil import tz
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference, Series
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# -------------------- Helpers --------------------

FILENAME_TS_RE = re.compile(r"_(\d{2})(\d{2})(\d{4})_(\d{2})_(\d{2})")  # _MMDDYYYY_HH_MM
DATE_PREFIX_RE = re.compile(r"^(\d{8})_")  # 8-digit date prefix

def load_config(path: Path) -> dict:
    if not path.exists():
        print(f"[ERROR] Config not found: {path}", file=sys.stderr)
        sys.exit(2)
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def parse_src_ts_from_filename(name: str):
    m = FILENAME_TS_RE.search(name)
    if not m:
        return None
    mm, dd, yyyy, HH, MM = m.groups()
    try:
        dt = datetime(int(yyyy), int(mm), int(dd), int(HH), int(MM))
        return dt
    except Exception:
        return None

def format_dt(dt: datetime, fmt: str, tzname: str = "America/New_York") -> str:
    # fmt tokens: DDMMYY_HH_mm and DDMMYYYY_HHmm per spec
    local = dt.astimezone(ZoneInfo(tzname))
    tokens = {
        "DD": f"{local.day:02d}",
        "MM": f"{local.month:02d}",
        "YY": f"{local.year % 100:02d}",
        "YYYY": f"{local.year:04d}",
        "HH": f"{local.hour:02d}",
        "mm": f"{local.minute:02d}",
    }
    out = fmt
    for k, v in tokens.items():
        out = out.replace(k, v)
    return out

def discover_latest_workbooks(paths, ignore_substrings, extensions):
    discovered = {}  # {ConoX_path: latest_file_path}
    for pattern in paths:
        for folder in Path().glob(pattern):
            if not folder.is_dir():
                continue
            # Identify Cono name from parent folder
            cono_name = folder.parent.name  # e.g., Cono1
            latest_file = None
            latest_key = None
            for f in folder.glob("**/*"):
                if not f.is_file():
                    continue
                if f.suffix.lower() not in extensions:
                    continue
                name_lower = f.name.lower()
                if any(sub.lower() in name_lower for sub in ignore_substrings):
                    continue
                # Prefer filename timestamp
                src_ts = parse_src_ts_from_filename(f.name)
                key = (1, src_ts) if src_ts else (0, f.stat().st_mtime)
                # Keep the max by tuple comparison; src_ts None falls back to mtime
                if latest_key is None or key > latest_key:
                    latest_key = key
                    latest_file = f
            if latest_file:
                discovered[cono_name] = latest_file
    return discovered

def find_section11_sheet_name(xl: pd.ExcelFile, candidates: list[str]) -> str | None:
    lower_sheets = {s.lower(): s for s in xl.sheet_names}
    for cand in candidates:
        # exact / close variants: compare lower without spaces/underscores/hyphens
        norm_cand = re.sub(r"[\s_-]+", "", cand.lower())
        for lo, orig in lower_sheets.items():
            if re.sub(r"[\s_-]+", "", lo) == norm_cand:
                return orig
    # fallback: any sheet that startswith "section11"
    for lo, orig in lower_sheets.items():
        if lo.replace(" ", "").startswith("section11"):
            return orig
    return None

def _coerce_numeric_df(df: pd.DataFrame) -> pd.DataFrame:
    for c in df.columns:
        if pd.api.types.is_object_dtype(df[c]):
            df[c] = pd.to_numeric(df[c], errors="ignore")
    return df

def section11_cleanup(xl_path: Path, config: dict) -> pd.DataFrame | None:
    sec11_cfg = config.get("section11", {})
    if not sec11_cfg.get("enabled", True):
        return None
    candidates = sec11_cfg.get("sheet_candidates", [])
    clusters = sec11_cfg.get("clusters", ["Login_Date", "LastWk"])

    try:
        xle = pd.ExcelFile(xl_path, engine="openpyxl")
    except Exception as e:
        print(f"[WARNING] Could not open for Section 11 cleanup: {xl_path.name}: {e}")
        return None

    sheet_name = find_section11_sheet_name(xle, candidates)
    if not sheet_name:
        print(f"[INFO] Section 11 sheet not found in {xl_path.name}")
        return None

    try:
        df = pd.read_excel(xle, sheet_name=sheet_name)
    except Exception as e:
        print(f"[WARNING] Failed reading Section 11 sheet '{sheet_name}': {e}")
        return None

    # Drop fully empty rows
    df = df.dropna(how="all")
    # Drop rows where both Oper and Vname blank (case-insensitively look for cols)
    cols_lower = {c.lower(): c for c in df.columns}
    oper_col = None
    vname_col = None
    for c in df.columns:
        cl = c.lower()
        if cl == "oper":
            oper_col = c
        if cl in ("vname", "v-name", "v_name", "v-name ", "v name"):
            vname_col = c
    if oper_col and vname_col:
        df = df[~(df[oper_col].isna() & df[vname_col].isna())]

    # Identify cluster columns by suffix patterns and date prefixes
    def pick_and_sort(suffix: str):
        cols = []
        for c in df.columns:
            if c.endswith("_" + suffix) or c.endswith(suffix):
                m = DATE_PREFIX_RE.match(c)
                if m:
                    date_key = m.group(1)  # yyyymmdd? actually MMDDYYYY per input; we'll treat as string order is fine because we'll parse
                    # Convert to comparable date
                    mm = date_key[0:2]; dd = date_key[2:4]; yyyy = date_key[4:8]
                    try:
                        dt = datetime(int(yyyy), int(mm), int(dd))
                    except Exception:
                        dt = None
                    cols.append((c, dt))
        # sort by date asc
        cols.sort(key=lambda x: (x[1] is None, x[1]))
        return [c for c, _ in cols]

    login_cols = pick_and_sort("Login_Date")
    lastwk_cols = pick_and_sort("LastWk")

    # Reorder columns: keep non-cluster first, then login cluster, then lastwk cluster
    non_cluster = [c for c in df.columns if c not in set(login_cols + lastwk_cols)]
    ordered = non_cluster + login_cols + lastwk_cols
    df = df.reindex(columns=ordered)

    # Highlight rule on rightmost vs left neighbor within each cluster
    df_clean = df.copy()

    # We'll return df_clean; highlighting is applied when writing with openpyxl
    # The function will attach metadata for highlighting
    df_clean.attrs["section11_highlight"] = {
        "login_cols": login_cols,
        "lastwk_cols": lastwk_cols
    }
    return df_clean

def ensure_min_rows(df: pd.DataFrame, min_rows: int) -> bool:
    if df is None:
        return False
    # Drop fully empty rows
    temp = df.dropna(how="all")
    return len(temp) >= min_rows

def gather_y_columns(df: pd.DataFrame, y_suffix: str) -> list[str]:
    cols = []
    for c in df.columns:
        if re.fullmatch(rf"\d{{8}}_{re.escape(y_suffix)}", str(c)):
            m = DATE_PREFIX_RE.match(str(c))
            if m:
                date_key = m.group(1)
                mm = date_key[0:2]; dd = date_key[2:4]; yyyy = date_key[4:8]
                try:
                    dt = datetime(int(yyyy), int(mm), int(dd))
                except Exception:
                    dt = None
                cols.append((c, dt))
    cols.sort(key=lambda x: (x[1] is None, x[1]))
    return [c for c, _ in cols]

def build_chart(ws_graph, chart_type, title, categories_ref, series_refs, series_titles):
    if chart_type == "line":
        ch = LineChart()
    elif chart_type in ("column", "stacked_column"):
        ch = BarChart()
        if chart_type == "stacked_column":
            ch.grouping = "stacked"
    elif chart_type == "bar":
        ch = BarChart()
        ch.type = "bar"
    elif chart_type == "scatter":
        ch = ScatterChart()
        ch.set_categories(categories_ref)
        for ref, stitle in zip(series_refs, series_titles):
            s = Series(ref, title=stitle)
            ch.series.append(s)
        ch.title = title
        ws_graph.add_chart(ch, "A1")
        return
    else:
        ch = LineChart()

    ch.title = title
    ch.set_categories(categories_ref)
    for ref, stitle in zip(series_refs, series_titles):
        s = Series(ref, title=stitle)
        ch.series.append(s)

    # Axis label rotation (openpyxl has limited support; applied to text properties)
    try:
        ch.x_axis.txPr = ch.x_axis.txPr or openpyxl.drawing.text.RichText()
        ch.x_axis.title = ""
        # Rotation hint not directly supported for category labels; rely on Excel's auto or user manual
    except Exception:
        pass

    ws_graph.add_chart(ch, "A1")

def write_dataframe(ws, df: pd.DataFrame, start_row=1, start_col=1):
    for j, col in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=j, value=col)
    for i, (_, row) in enumerate(df.iterrows(), start=start_row+1):
        for j, col in enumerate(df.columns, start=start_col):
            ws.cell(row=i, column=j, value=row[col])

def apply_section11_highlighting(ws, df: pd.DataFrame, start_row=1, start_col=1):
    meta = df.attrs.get("section11_highlight", {})
    login_cols = meta.get("login_cols", [])
    lastwk_cols = meta.get("lastwk_cols", [])
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    def highlight_cluster(cols):
        if len(cols) < 2:
            return
        # Locate columns indexes in the written sheet (1-based)
        header_to_colidx = {}
        for j, col in enumerate(df.columns, start=start_col):
            header_to_colidx[col] = j
        right = cols[-1]
        left = cols[-2]
        if right not in header_to_colidx or left not in header_to_colidx:
            return
        c_right = header_to_colidx[right]
        c_left = header_to_colidx[left]
        # For each data row
        for r in range(start_row+1, start_row+1+len(df.index)):
            v_right = ws.cell(row=r, column=c_right).value
            v_left = ws.cell(row=r, column=c_left).value
            try:
                if v_right == v_left and v_right is not None:
                    ws.cell(row=r, column=c_right).fill = highlight_fill
            except Exception:
                pass

    highlight_cluster(login_cols)
    highlight_cluster(lastwk_cols)

def reorder_sheets(wb: Workbook, graph_names: list[str], modified_names: list[str]):
    # Collect worksheets in desired order: graphs first, then modified, then the rest (originals)
    desired_order = []
    name_to_sheet = {ws.title: ws for ws in wb.worksheets}
    for n in graph_names:
        if n in name_to_sheet:
            desired_order.append(name_to_sheet[n])
    for n in modified_names:
        if n in name_to_sheet:
            desired_order.append(name_to_sheet[n])
    # originals: any not already in desired_order
    already = set(ws.title for ws in desired_order)
    for ws in wb.worksheets:
        if ws.title not in already:
            desired_order.append(ws)
    wb._sheets = desired_order  # openpyxl internal but acceptable

def append_run_log(run_info: str, tzname="America/New_York", log_path=Path("graph_generation_log.txt")):
    dt = datetime.now(ZoneInfo(tzname))
    stamp = dt.strftime("%m%d%Y_%H%M")
    with log_path.open("a", encoding="utf-8") as f:
        f.write(f"[{stamp}] {run_info}\n")

# -------------------- Main process --------------------

def process_workbook(cono: str, src_path: Path, cfg: dict, dryrun=False):
    print(f"[INFO] Processing workbook for {cono}: {src_path}")
    apply_all = cfg.get("workbooks", [])[0] if cfg.get("workbooks") else {}
    charts = apply_all.get("charts", [])
    rules = cfg.get("rules", {})
    min_rows = int(rules.get("min_rows_per_sheet", 2))
    tzname = cfg.get("output", {}).get("timezone", "America/New_York")
    name_collision_suffix = cfg.get("output", {}).get("name_collision_suffix_for_original", " (orig)")

    # Read all sheets quickly to pandas once
    try:
        xle = pd.ExcelFile(src_path, engine="openpyxl")
    except Exception as e:
        print(f"[ERROR] Cannot open {src_path.name}: {e}")
        return None

    # Build modified sheets (Section11_Clean)
    modified_sheets = {}
    sec11_df = section11_cleanup(src_path, cfg)
    if sec11_df is not None and ensure_min_rows(sec11_df, 1):
        modified_sheets[cfg["section11"]["output_sheet"]] = sec11_df
        sec11_status = "cleaned/created"
    else:
        sec11_status = "not found/empty"
    print(f"[INFO] Section 11 status: {sec11_status}")

    # Prepare output workbook (we'll write data first, charts later)
    wb = Workbook()
    # remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    used_names = set()

    # Write modified sheets first (data; we'll reorder later)
    for name, df in modified_sheets.items():
        wsname = name
        if wsname in used_names:
            wsname += name_collision_suffix
        ws = wb.create_sheet(wsname)
        write_dataframe(ws, df)
        apply_section11_highlighting(ws, df)
        used_names.add(wsname)

    # Write original sheets (data)
    for sheet_name in xle.sheet_names:
        try:
            df = pd.read_excel(xle, sheet_name=sheet_name)
        except Exception as e:
            print(f"[WARNING] Skipping original sheet {sheet_name}: {e}")
            continue
        wsname = sheet_name
        if wsname in used_names:
            wsname += name_collision_suffix
        ws = wb.create_sheet(wsname)
        write_dataframe(ws, df)
        used_names.add(wsname)

    # Build charts -> each chart in its own sheet referencing data sheets we just wrote
    graph_sheet_names = []
    for ch in charts:
        src_sheet = ch.get("sheet")
        x_col = ch.get("x_col")
        y_suffix = ch.get("y_suffix")
        chart_type = ch.get("chart_type", cfg["charts_defaults"]["chart_type"])
        out_sheet = ch.get("output_sheet", f"Graph_{src_sheet}")
        title = ch.get("title", out_sheet)

        # Fetch data frame from output workbook (we wrote originals with same names unless collision occurred)
        target_sheet_name = src_sheet
        if target_sheet_name not in used_names and f"{src_sheet}{name_collision_suffix}" in used_names:
            target_sheet_name = f"{src_sheet}{name_collision_suffix}"

        if target_sheet_name not in [ws.title for ws in wb.worksheets]:
            print(f"[WARNING] Chart skipped (missing sheet): {src_sheet}")
            continue

        # Recreate df from the workbook sheet
        ws_data = wb[target_sheet_name]
        # Read header
        headers = [cell.value for cell in ws_data[1]]
        if x_col not in headers:
            print(f"[WARNING] Chart skipped (missing x_col '{x_col}') in sheet {src_sheet}")
            continue
        # Build a pandas dataframe from the sheet quickly
        data = []
        for r in ws_data.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in r):
                continue
            data.append(r)
        df = pd.DataFrame(data, columns=headers)

        if not ensure_min_rows(df, min_rows):
            print(f"[WARNING] Chart skipped (<{min_rows} rows) for sheet {src_sheet}")
            continue

        y_cols = gather_y_columns(df, y_suffix)
        if not y_cols:
            print(f"[WARNING] Chart skipped (no columns matched y_suffix '{y_suffix}') for sheet {src_sheet}")
            continue

        # Drop rows where all Y are empty
        if cfg.get("rules", {}).get("drop_rows_with_no_y", True):
            mask = df[y_cols].notna().any(axis=1)
            df = df.loc[mask]

        if not ensure_min_rows(df, min_rows):
            print(f"[WARNING] Chart skipped after cleaning (<{min_rows} rows) for sheet {src_sheet}")
            continue

        # Coerce numeric for y columns
        for c in y_cols:
            df[c] = pd.to_numeric(df[c], errors="coerce")

        # Write a temporary data table on the graph sheet for openpyxl references
        ws_graph = wb.create_sheet(out_sheet)
        graph_sheet_names.append(out_sheet)

        # Compose a small table: first col = x_col, then each y_col
        table_cols = [x_col] + y_cols
        # headers in friendly form for series titles
        series_titles = []
        for c in y_cols:
            m = DATE_PREFIX_RE.match(str(c))
            label = str(c)
            if m:
                date_key = m.group(1)
                mm = date_key[0:2]; dd = date_key[2:4]; yyyy = date_key[4:8]
                try:
                    dt = datetime(int(yyyy), int(mm), int(dd))
                    label = dt.strftime("%m/%d/%Y")
                except Exception:
                    pass
            series_titles.append(label)

        df_small = df[table_cols].copy()
        # Write table
        write_dataframe(ws_graph, df_small, start_row=1, start_col=1)

        max_row = 1 + len(df_small.index)
        max_col = len(table_cols)
        cats = Reference(ws_graph, min_col=1, min_row=2, max_col=1, max_row=max_row)
        series_refs = [
            Reference(ws_graph, min_col=idx+1, min_row=1, max_col=idx+1, max_row=max_row)
            for idx in range(1, max_col)
        ]

        # Build and add chart
        build_chart(ws_graph, chart_type, title, cats, series_refs, series_titles)
        print(f"[INFO] Chart created: {out_sheet} (from {src_sheet}, y_suffix={y_suffix})")

    # Reorder sheets: graphs first, then modified, then originals
    reorder_sheets(wb, graph_sheet_names, list(modified_sheets.keys()))

    if dryrun:
        print(f"[DRYRUN] Would write output workbook for {cono}")
        return None

    # Compute output filename
    src_ts = parse_src_ts_from_filename(src_path.name) or datetime.fromtimestamp(src_path.stat().st_mtime)
    run_ts = datetime.now(ZoneInfo(tzname))
    pattern = cfg.get("output", {}).get("filename_pattern")
    # Tokens: {cono}, {src_ts:DDMMYY_HH_mm}, {run_ts_EST:DDMMYYYY_HHmm}
    out_name = pattern
    out_name = out_name.replace("{cono}", cono)
    out_name = out_name.replace("{src_ts:DDMMYY_HH_mm}", format_dt(src_ts, "DDMMYY_HH_mm", tzname))
    out_name = out_name.replace("{run_ts_EST:DDMMYYYY_HHmm}", format_dt(run_ts, "DDMMYYYY_HHmm", tzname))
    out_name += ".xlsx" if not out_name.endswith(".xlsx") else ""

    out_path = src_path.parent / out_name
    wb.save(out_path)
    print(f"[SUCCESS] Wrote: {out_path}")
    return out_path

def main():
    parser = argparse.ArgumentParser(description="Generate graphs and Section 11 cleanup per config.")
    parser.add_argument("--config", default="config_grph.json", help="Path to config JSON (default: config_grph.json)")
    parser.add_argument("--workbooks", default="./Cono*/Consolidated_reports", help="Glob for ConoX Consolidated_reports folders")
    parser.add_argument("--dryrun", action="store_true", help="If set, do not write output files")
    args = parser.parse_args()

    cfg = load_config(Path(args.config))

    # Discover latest workbooks
    discovered = discover_latest_workbooks(
        paths=cfg.get("discovery", {}).get("paths", [args.workbooks]),
        ignore_substrings=cfg.get("discovery", {}).get("ignore_filename_contains", []),
        extensions=set(cfg.get("discovery", {}).get("file_extensions", [".xlsx"]))
    )
    if not discovered:
        print("[WARNING] No workbooks discovered.")
        append_run_log("No workbooks discovered.")
        sys.exit(0)

    # Process each discovered workbook
    outputs = []
    for cono, src_path in discovered.items():
        out = process_workbook(cono, src_path, cfg, dryrun=args.dryrun)
        outputs.append((cono, src_path, out))

    # Append run log
    lines = []
    for cono, src, out in outputs:
        status = "DRYRUN" if args.dryrun else ("SUCCESS" if out else "SKIPPED/ERROR")
        lines.append(f"{status} | {cono} | src={src} | out={out}")
    append_run_log(" ; ".join(lines))

if __name__ == "__main__":
    main()